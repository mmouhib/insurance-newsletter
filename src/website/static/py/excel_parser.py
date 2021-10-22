import openpyxl


def excel(path):
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    emails = []
    for i in range(1, sheet_obj.max_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)
        emails.append(str(cell_obj.value))
    return emails
